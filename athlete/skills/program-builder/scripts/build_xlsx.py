#!/usr/bin/env python3
"""
build_xlsx.py - render a training programme into a colour-banded Excel workbook.

Input: a JSON file containing the programme rows, either as
  (a) {"rows": [[14 fields], ...]}  with fields in the exact column order, or
  (b) a bare list  [[14 fields], ...].
A 13-field row predates the Category column and is padded; see rows_common.

Loading, validation, week numbers and day ordering come from rows_common, so this
script and build_program_json.py read the same rows.json identically. The two
outputs describe the same programme in the same order, or neither is produced.

ONE SHEET PER WEEK
------------------
rows.json carries every week of the block (see build_program_json.py), so the
workbook gets one tab per week - "Week 1", "Week 2", ... - in numeric order. A
single-week rows.json simply produces a single tab.

Each training day (parsed from the "Day" column, e.g. "Day 1 (Mon) - ...") gets
its own background fill band, with a medium top-border divider where the day
changes.

Usage:
  python3 build_xlsx.py --input rows.json --output "Block - Programme.xlsx"
  python3 build_xlsx.py --input rows.json --output "..." --title "Wk"   # tab prefix
"""
import argparse
import os
import sys

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from rows_common import (COLS, DAY, die, day_order, day_slots,  # noqa: E402
                         load_rows, sort_rows, week_of)

# Day band palette (light fills); extends by cycling if a block has >6 days.
DAY_BANDS = ["DDEBF7", "E2EFDA", "FCE4D6", "EDE7F6", "FFF2CC", "E7E6E6"]
DAY_LABEL = ["1F4E79", "375623", "833C00", "4B2E83", "7F6000", "3B3838"]

WRAP_COLS = {"Load", "Completed Notes", "Focus / Notes", "Progression Rule"}
WIDTHS = {"Week": 6, "Day": 30, "Exercise": 26, "Sets": 9, "Reps": 22, "Load": 26,
          "Intensity (RPE)": 18, "Tempo": 9, "Rest": 14, "Completed": 11,
          "Completed Notes": 34, "Focus / Notes": 62, "Progression Rule": 58,
          "Category": 12}

MAX_TAB = 31  # Excel's hard limit on a worksheet name


def write_sheet(ws, rows, band_of):
    header_fill = PatternFill("solid", fgColor="1F4E5F")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    base_font = Font(name="Arial", size=10)
    thin = Side(style="thin", color="BFBFBF")
    medium_top = Side(style="medium", color="808080")

    ws.append(COLS)
    for c in ws[1]:
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for r in rows:
        ws.append(r)

    prev_day = None
    for ridx in range(2, ws.max_row + 1):
        day_val = ws.cell(ridx, DAY + 1).value or ""
        di = band_of[day_val]
        band = DAY_BANDS[di % len(DAY_BANDS)]
        label_color = DAY_LABEL[di % len(DAY_LABEL)]
        is_start = day_val != prev_day
        for cidx in range(1, len(COLS) + 1):
            cell = ws.cell(ridx, cidx)
            header = COLS[cidx - 1]
            cell.font = base_font
            cell.fill = PatternFill("solid", fgColor=band)
            top = medium_top if is_start else thin
            cell.border = Border(left=thin, right=thin, top=top, bottom=thin)
            wrap = header in WRAP_COLS
            cell.alignment = Alignment(
                vertical="top", wrap_text=wrap,
                horizontal="left" if wrap or header in ("Day", "Exercise") else "center")
        ws.cell(ridx, DAY + 1).font = Font(name="Arial", size=10, bold=True,
                                           color=label_color)
        prev_day = day_val

    for name, w in WIDTHS.items():
        ws.column_dimensions[get_column_letter(COLS.index(name) + 1)].width = w
    ws.freeze_panes = "A2"
    ws.sheet_view.showGridLines = False


def build(rows, out_path, title="Week"):
    order = day_order(rows)
    # Same slot numbering as the exercise ids in program.json, so a day's colour
    # here and its 'd<n>' there always agree.
    band_of = {d: n - 1 for d, n in day_slots(order).items()}

    by_week = {}
    for r in sort_rows(rows, order):
        by_week.setdefault(week_of(r), []).append(r)

    tabs = {w: f"{title} {w}"[:MAX_TAB] for w in sorted(by_week)}
    if len(set(tabs.values())) != len(tabs):
        die(f"--title {title!r} is too long: week tabs collide once truncated to "
            f"{MAX_TAB} characters. Use a shorter prefix.")

    wb = Workbook()
    wb.remove(wb.active)
    for week in sorted(by_week):
        write_sheet(wb.create_sheet(title=tabs[week]), by_week[week], band_of)
    wb.save(out_path)

    print(f"OK: {len(rows)} rows across {len(by_week)} week-sheet(s) -> {out_path}")
    for week in sorted(by_week):
        days = {}
        for r in by_week[week]:
            k = r[DAY].split(" - ")[0]
            days[k] = days.get(k, 0) + 1
        print(f"  {tabs[week]}: {len(by_week[week])} rows; "
              + "; ".join(f"{k}={v}" for k, v in days.items()))


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True, help="JSON file of programme rows")
    ap.add_argument("--output", required=True, help="output .xlsx path")
    ap.add_argument("--title", default="Week",
                    help="worksheet tab prefix; the week number is appended")
    args = ap.parse_args()
    build(load_rows(args.input), args.output, args.title)


if __name__ == "__main__":
    main()
